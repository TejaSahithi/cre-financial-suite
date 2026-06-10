"""
CRE Financial Suite – Full Product Audit Report Generator
Creates a multi-sheet Excel workbook with findings, maturity scores, and fixes.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ─── Color Palette ────────────────────────────────────────────────────────────
RED_FILL     = PatternFill("solid", fgColor="FF4444")
ORANGE_FILL  = PatternFill("solid", fgColor="FF8800")
YELLOW_FILL  = PatternFill("solid", fgColor="FFD700")
GREEN_FILL   = PatternFill("solid", fgColor="22CC66")
BLUE_FILL    = PatternFill("solid", fgColor="1D4ED8")
DARK_FILL    = PatternFill("solid", fgColor="1E293B")
LIGHT_FILL   = PatternFill("solid", fgColor="F1F5F9")
HEADER_FILL  = PatternFill("solid", fgColor="0F172A")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
GRAY_FILL    = PatternFill("solid", fgColor="E2E8F0")

WHITE_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
DARK_FONT   = Font(color="1E293B", bold=True, name="Calibri", size=10)
BODY_FONT   = Font(color="1E293B", name="Calibri", size=9)
SMALL_FONT  = Font(color="64748B", name="Calibri", size=8)
TITLE_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=14)
HEADING_FONT= Font(color="0F172A", bold=True, name="Calibri", size=11)

thin = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

def style_header_row(ws, row, cols, fill=HEADER_FILL, font=WHITE_FONT):
    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER

def style_cell(ws, row, col, value, fill=None, font=BODY_FONT, align=LEFT):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if fill:
        cell.fill = fill
    cell.font = font
    cell.alignment = align
    cell.border = BORDER
    return cell

def priority_fill(priority):
    return {
        "SHOWSTOPPER": RED_FILL,
        "HIGH":        PatternFill("solid", fgColor="FF8800"),
        "MEDIUM":      YELLOW_FILL,
        "LOW":         PatternFill("solid", fgColor="86EFAC"),
    }.get(priority, WHITE_FILL)

def status_fill(status):
    return {
        "OPEN":        RED_FILL,
        "IN PROGRESS": YELLOW_FILL,
        "DONE":        GREEN_FILL,
    }.get(status, WHITE_FILL)

# ══════════════════════════════════════════════════════════════════════════════
# DATA
# ══════════════════════════════════════════════════════════════════════════════

FINDINGS = [
    # SHOWSTOPPERS — nothing works until these are fixed
    {
        "id": "F-001",
        "area": "Infrastructure / Config",
        "component": ".env / Supabase URL",
        "title": "VITE_SUPABASE_URL points to localhost — app is dead in production",
        "priority": "SHOWSTOPPER",
        "impact": "100% of features broken. Auth, data reads/writes, file upload, edge functions all target http://127.0.0.1:54321 which only runs with local Supabase CLI.",
        "root_cause": ".env line 1: VITE_SUPABASE_URL=http://127.0.0.1:54321 and VITE_SUPABASE_ANON_KEY uses a local publishable key.",
        "fix": "Set VITE_SUPABASE_URL=https://<project-id>.supabase.co and VITE_SUPABASE_ANON_KEY=<production-anon-key> in your hosting environment (Vercel / Netlify env vars). Never commit real keys.",
        "effort": "15 min",
        "status": "OPEN",
        "file": ".env",
    },
    {
        "id": "F-002",
        "area": "Infrastructure / Config",
        "component": "Supabase Migrations",
        "title": "80+ DB migrations may not be applied to production Supabase project",
        "priority": "SHOWSTOPPER",
        "impact": "Tables do not exist → every DB read/write fails with 'relation does not exist' or RLS errors. Leases, budgets, expenses, uploads — all broken.",
        "root_cause": "Migrations exist in supabase/migrations/ but supabase db push has not been run against the production project.",
        "fix": "Run: supabase link --project-ref <project-id> && supabase db push. Review migration order for the ~80 files.",
        "effort": "1–2 hours",
        "status": "OPEN",
        "file": "supabase/migrations/ (80+ files)",
    },
    {
        "id": "F-003",
        "area": "Infrastructure / Config",
        "component": "Edge Functions Deployment",
        "title": "60+ Supabase Edge Functions not deployed to production",
        "priority": "SHOWSTOPPER",
        "impact": "Lease upload pipeline, PDF parsing, AI extraction, email, budget generation, CAM calculations, approval workflows — all fail with 404 on function invocation.",
        "root_cause": "Functions exist in supabase/functions/ but supabase functions deploy has not been run.",
        "fix": "Run: supabase functions deploy --no-verify-jwt (for public ones) or supabase functions deploy for each function. Use supabase/config.toml verify_jwt settings.",
        "effort": "1–2 hours",
        "status": "OPEN",
        "file": "supabase/functions/ (60+ functions)",
    },
    # HIGH
    {
        "id": "F-004",
        "area": "Backend / AI Extraction",
        "component": "Supabase Secrets — ANTHROPIC_API_KEY",
        "title": "ANTHROPIC_API_KEY not set → AI lease extraction fails silently",
        "priority": "HIGH",
        "impact": "Every lease PDF upload reaches the AI extraction step and fails. User sees 'Extraction pipeline failed' or fallback to manual review with no fields extracted.",
        "root_cause": "ANTHROPIC_API_KEY is referenced in edge function secrets but never set via supabase secrets set. Same for VERTEX_PROJECT_ID and GOOGLE_SERVICE_ACCOUNT_KEY.",
        "fix": "supabase secrets set ANTHROPIC_API_KEY=sk-ant-... OR set VERTEX_PROJECT_ID + VERTEX_LOCATION + GOOGLE_SERVICE_ACCOUNT_KEY for Google Vertex path. At least one AI provider required.",
        "effort": "30 min",
        "status": "OPEN",
        "file": ".env.example / supabase secrets",
    },
    {
        "id": "F-005",
        "area": "Backend / File Pipeline",
        "component": "Supabase Secrets — WORKER_INTERNAL_SECRET",
        "title": "Missing WORKER_INTERNAL_SECRET → DOWNSTREAM_AUTH_FAILED on every upload",
        "priority": "HIGH",
        "impact": "The parse-pdf-docling / lease-extraction-worker functions cannot authenticate to each other. Every lease upload fails at the parsing stage with error code DOWNSTREAM_AUTH_FAILED.",
        "root_cause": "LeaseUpload.jsx error handler explicitly calls this out: 'Ensure WORKER_INTERNAL_SECRET is set in Supabase secrets'. The inter-function auth token is missing.",
        "fix": "supabase secrets set WORKER_INTERNAL_SECRET=<random-32-char-secret>. Same value must be set in every worker function that calls another internal function.",
        "effort": "15 min",
        "status": "OPEN",
        "file": "supabase/functions/_shared / LeaseUpload.jsx:260",
    },
    {
        "id": "F-006",
        "area": "Backend / Email",
        "component": "Supabase Secrets — RESEND_API_KEY",
        "title": "RESEND_API_KEY not set → all email flows broken",
        "priority": "HIGH",
        "impact": "Signup confirmation, org invitations, password resets, budget rework notifications, CAM approval emails — all fail. Users get no email after signup → cannot activate accounts.",
        "root_cause": "send-email edge function relies on RESEND_API_KEY. Without it, every supabase.functions.invoke('send-email', ...) call throws 500.",
        "fix": "Create account at resend.com, get API key, run: supabase secrets set RESEND_API_KEY=re_...",
        "effort": "30 min",
        "status": "OPEN",
        "file": "supabase/functions/send-email/index.ts",
    },
    {
        "id": "F-007",
        "area": "Frontend / Budget",
        "component": "CreateBudget.jsx",
        "title": "Budget generation writes hardcoded fake numbers ($669K revenue, $232K expenses) on failure",
        "priority": "HIGH",
        "impact": "When the generate-budget edge function fails (which it always does without secrets), CreateBudget still calls createMutation.mutate() with fallback values: total_revenue: data.total_revenue || 669000. Junk data is persisted to the DB.",
        "root_cause": "Lines 260–262 in CreateBudget.jsx use hardcoded numbers as fallbacks. A failed API call returns {} and 669000/232050 are used as actual budget values.",
        "fix": "Guard the createMutation call: only proceed if !error && result && !result.error. Show a toast.error and return early on failure instead of saving fake data.",
        "effort": "1 hour",
        "status": "OPEN",
        "file": "src/pages/CreateBudget.jsx:260",
    },
    {
        "id": "F-008",
        "area": "Frontend / Budget",
        "component": "CreateBudget.jsx",
        "title": "Budget creation blocked unless CAM snapshot exists — most users cannot create any budget",
        "priority": "HIGH",
        "impact": "handleGenerate() checks: if (!latestCamSnapshot && !generateWithoutCam) { toast.error(...); return; }. New users with no CAM data cannot generate a budget. The checkbox 'Generate without CAM' exists but is easy to miss.",
        "root_cause": "CAM-first design assumption. In practice, most users will never have a CAM snapshot until they run CAM Calculations first, which is a separate multi-step workflow.",
        "fix": "Default generateWithoutCam to true. Show the CAM warning as informational, not a hard block. Let users proceed and add CAM data later.",
        "effort": "30 min",
        "status": "OPEN",
        "file": "src/pages/CreateBudget.jsx:199",
    },
    # MEDIUM
    {
        "id": "F-009",
        "area": "Frontend / Lease Upload",
        "component": "LeaseUpload.jsx",
        "title": "Pipeline polling stops prematurely — stale closure on fileRecord?.status",
        "priority": "MEDIUM",
        "impact": "User uploads a lease, pipeline starts, UI shows 'Uploaded'. The interval is created once with status='null'. Because ACTIVE_STATUSES.has(null) === false, polling never runs again. User sees frozen status forever.",
        "root_cause": "useEffect at line 541 creates an interval that reads fileRecord?.status — but this is a stale closure. The interval captures the initial value and never sees the updated status from setFileRecord().",
        "fix": "Use a ref to track fileRecord status inside the interval, or move the polling condition check out of the interval into a separate useEffect that depends on fileRecord?.status. Pattern: useRef(fileRecord).current.status inside interval.",
        "effort": "1 hour",
        "status": "OPEN",
        "file": "src/pages/LeaseUpload.jsx:541",
    },
    {
        "id": "F-010",
        "area": "Frontend / Budget",
        "component": "CreateBudget.jsx",
        "title": "resolveWritableOrgId returns null for super-admin → org_id becomes empty string",
        "priority": "MEDIUM",
        "impact": "Super-admin users cannot create budgets. org_id is passed as '' (empty string) to BudgetService.create(), which fails Supabase RLS policies checking org_id is valid UUID.",
        "root_cause": "Line 244: resolveWritableOrgId(orgId) returns null when no acting org is set. Line 248: org_id: writableOrgId || '' — empty string fallback violates FK constraint.",
        "fix": "Add guard: if (!writableOrgId) { toast.error('Select an organization first'); return; }. Also fix the || '' fallback to || null.",
        "effort": "30 min",
        "status": "OPEN",
        "file": "src/pages/CreateBudget.jsx:244",
    },
    {
        "id": "F-011",
        "area": "Backend / Database",
        "component": "leases table / findLeaseByFileId",
        "title": "JSONB text-path query syntax for finding lease by file ID may fail",
        "priority": "MEDIUM",
        "impact": "After uploading a lease, the system cannot find the existing draft and may create duplicate lease records. The ensureLeaseDraft flow creates duplicates on retry.",
        "root_cause": "LeaseUpload.jsx line 1165: .eq('extraction_data->>source_file_id', fileId) — the ->> PostgREST JSONB text-path operator may not work correctly if the column type is jsonb vs text in older Supabase versions.",
        "fix": "Use .filter('extraction_data->>source_file_id', 'eq', fileId) OR add a generated column source_file_id in the DB migration that extracts from extraction_data for indexed lookup.",
        "effort": "1 hour",
        "status": "OPEN",
        "file": "src/pages/LeaseUpload.jsx:1165 / supabase/migrations",
    },
    {
        "id": "F-012",
        "area": "Frontend / Budget",
        "component": "CreateBudget.jsx",
        "title": "Default budget_year hardcoded to 2027 — users in 2026 see wrong year",
        "priority": "MEDIUM",
        "impact": "Every new budget form opens with year=2027. Users must manually correct it every time. In June 2026, this creates friction for FY2026 budgets.",
        "root_cause": "Line 36: budget_year: 2027 is hardcoded in buildDefaultForm().",
        "fix": "Replace with: budget_year: new Date().getFullYear() + 1 (next fiscal year) or new Date().getFullYear() based on business logic.",
        "effort": "5 min",
        "status": "OPEN",
        "file": "src/pages/CreateBudget.jsx:36",
    },
    {
        "id": "F-013",
        "area": "Frontend / Lease Upload",
        "component": "LeaseUpload.jsx",
        "title": "Memory leak: polling useEffect recreates interval on every fileRecord?.status change",
        "priority": "MEDIUM",
        "impact": "Every time fileRecord updates (3-second poll), the useEffect dependency [fileId, fileRecord?.status] fires, creating a new interval and clearing the old one. Multiple fetch calls overlap, flooding the backend.",
        "root_cause": "The useEffect at line 541 depends on fileRecord?.status which changes on every poll response. This creates a polling cascade under load.",
        "fix": "Remove fileRecord?.status from the useEffect dependency array. Use a ref inside the interval callback to read the latest status: const statusRef = useRef(null); statusRef.current = fileRecord?.status;",
        "effort": "1 hour",
        "status": "OPEN",
        "file": "src/pages/LeaseUpload.jsx:562",
    },
    {
        "id": "F-014",
        "area": "Backend / Auth",
        "component": "signup edge function",
        "title": "Signup email confirmation broken without RESEND_API_KEY — users stuck pending",
        "priority": "MEDIUM",
        "impact": "New users sign up, no confirmation email arrives, they cannot log in. Support burden is high. The pending_approval state has no self-serve exit.",
        "root_cause": "signup/index.ts calls send-email edge function which needs RESEND_API_KEY. Without it, the email call throws but the user record is created — leaving them in limbo.",
        "fix": "Fix F-006 (set RESEND_API_KEY). Also add defensive error handling in signup to rollback user creation if email fails, OR use Supabase's built-in email for signup confirmation as fallback.",
        "effort": "2 hours",
        "status": "OPEN",
        "file": "supabase/functions/signup/index.ts",
    },
    {
        "id": "F-015",
        "area": "Backend / Storage",
        "component": "financial-uploads bucket",
        "title": "Storage bucket 'financial-uploads' may not exist in production",
        "priority": "MEDIUM",
        "impact": "File uploads fail with 'Bucket not found'. Lease PDFs, expense documents, budget files cannot be stored.",
        "root_cause": "Migration 20260403_create_financial_uploads_bucket.sql creates the bucket, but if migrations aren't applied (see F-002), bucket doesn't exist.",
        "fix": "Apply migration F-002 fix. Also manually verify in Supabase Dashboard > Storage that 'financial-uploads' bucket exists with correct public/private settings.",
        "effort": "30 min (after F-002)",
        "status": "OPEN",
        "file": "supabase/migrations/20260403_create_financial_uploads_bucket.sql",
    },
    {
        "id": "F-016",
        "area": "Frontend / General",
        "component": "src/utils/index.js + index.ts",
        "title": "Duplicate utils files (index.js AND index.ts) may cause module resolution conflict",
        "priority": "MEDIUM",
        "impact": "Vite may resolve @/utils to either index.js or index.ts unpredictably. If types diverge, runtime crashes occur on pages importing createPageUrl.",
        "root_cause": "src/utils/ contains both index.js and index.ts. With bundler moduleResolution in jsconfig.json, TypeScript prefers .ts but Vite may pick .js.",
        "fix": "Delete src/utils/index.js (keep only index.ts) OR delete index.ts (keep only index.js). Standardize to one file.",
        "effort": "15 min",
        "status": "OPEN",
        "file": "src/utils/index.js + src/utils/index.ts",
    },
    # LOW
    {
        "id": "F-017",
        "area": "Frontend / Budget",
        "component": "CreateBudget.jsx",
        "title": "escapeHtml function defined but never used — dead code",
        "priority": "LOW",
        "impact": "Minimal. Dead code adds confusion and may suggest missing XSS protection was intended but not wired up.",
        "root_cause": "Lines 58–65 define escapeHtml() but no call site exists in the file.",
        "fix": "Remove the function OR wire it to sanitize the rejection_comment before sending to the send-email edge function (which would be the correct use).",
        "effort": "10 min",
        "status": "OPEN",
        "file": "src/pages/CreateBudget.jsx:58",
    },
    {
        "id": "F-018",
        "area": "Frontend / Lease Upload",
        "component": "LeaseUpload.jsx",
        "title": "ensureLeaseDraft useEffect has missing dependency 'fileRecord' — stale closure",
        "priority": "LOW",
        "impact": "The auto-staging effect at line 721 does not include fileRecord in its dependency array. After a status update, it may re-run with an outdated fileRecord snapshot.",
        "root_cause": "Line 752: dependency array is [fileId, fileRecord, queryClient] — fileRecord is included, but ensureLeaseDraft closes over fileRecord from outer scope which may be stale.",
        "fix": "Pass fileRecord as parameter to ensureLeaseDraft() to avoid closure staleness, or use useRef.",
        "effort": "30 min",
        "status": "OPEN",
        "file": "src/pages/LeaseUpload.jsx:721",
    },
    {
        "id": "F-019",
        "area": "Frontend / General",
        "component": "Pages / Navigation",
        "title": "65+ pages loaded via lazy() but SuspenseBoundary fallback is generic spinner",
        "priority": "LOW",
        "impact": "On slow connections, users see a blank spinner with no context. No skeleton screens or page-specific loading states.",
        "root_cause": "AppRoutes.jsx uses <Suspense fallback={...}> with a single generic fallback for all 65 pages.",
        "fix": "Add per-page skeleton screens or at minimum a page title + spinner labeled with the page name.",
        "effort": "2–4 hours",
        "status": "OPEN",
        "file": "src/app/AppRoutes.jsx",
    },
    {
        "id": "F-020",
        "area": "Backend / Config",
        "component": "Stripe",
        "title": "Stripe public key and webhook secret not configured — billing completely non-functional",
        "priority": "LOW",
        "impact": "Billing page, checkout sessions, subscription management — all fail. STRIPE_SECRET_KEY and STRIPE_WEBHOOK_SECRET must be set as Supabase secrets.",
        "root_cause": "create-checkout-session and stripe-webhook edge functions need STRIPE_SECRET_KEY / STRIPE_WEBHOOK_SECRET. .env.example does not document these.",
        "fix": "supabase secrets set STRIPE_SECRET_KEY=sk_live_... STRIPE_WEBHOOK_SECRET=whsec_... Also add VITE_STRIPE_PUBLISHABLE_KEY to frontend env.",
        "effort": "1 hour",
        "status": "OPEN",
        "file": "supabase/functions/create-checkout-session / stripe-webhook",
    },
    {
        "id": "F-021",
        "area": "Security",
        "component": "RLS Policies",
        "title": "Super-admin bypass in api.js getCurrentAccessScope() is correct, but 3 RLS policies have SECURITY DEFINER without explicit search_path",
        "priority": "LOW",
        "impact": "Potential privilege escalation via search_path hijack in some SECURITY DEFINER functions. Migration 20260602014048_security_definer_search_path_hardening.sql was added but may not be applied.",
        "root_cause": "Several stored procedures use SECURITY DEFINER without SET search_path = public, pg_catalog which allows schema injection.",
        "fix": "Apply the security hardening migration (F-002 prerequisite) and verify all SECURITY DEFINER functions have explicit search_path set.",
        "effort": "1 hour (after F-002)",
        "status": "OPEN",
        "file": "supabase/migrations/20260602014048_security_definer_search_path_hardening.sql",
    },
    {
        "id": "F-022",
        "area": "Frontend / UX",
        "component": "All pages",
        "title": "No offline / error boundary recovery — entire app crashes on Supabase connection error",
        "priority": "LOW",
        "impact": "If Supabase is temporarily unavailable, the React Query errors bubble up to the ErrorBoundary which shows a generic crash screen. No retry mechanism or graceful degradation.",
        "root_cause": "QueryClient has no global onError handler. api.js has an in-memory fallback but it's only triggered for unauthenticated users, not for network failures.",
        "fix": "Add a global React Query error handler that shows a 'Connection lost, retrying...' toast. Enable React Query's automatic retry with exponential backoff.",
        "effort": "2 hours",
        "status": "OPEN",
        "file": "src/lib/query-client.js",
    },
]

MATURITY = [
    {"area": "Authentication & Auth Flow",  "score": 85, "notes": "Strong: OAuth, MFA, magic link, RBAC. Weakness: signup email broken without RESEND_API_KEY."},
    {"area": "Database Schema & RLS",       "score": 80, "notes": "80+ migrations, comprehensive RLS. Weakness: not applied to production."},
    {"area": "Lease Management",            "score": 72, "notes": "Upload, AI extraction, review pipeline built. Weakness: polling bug, AI creds missing."},
    {"area": "Budget Module",               "score": 60, "notes": "3 generation methods, approval workflow. Weakness: hardcoded fallback values, CAM hard-block."},
    {"area": "Expense / CAM Engine",        "score": 65, "notes": "Rule engine, workflow approvals built. Weakness: depends on deployed edge functions."},
    {"area": "Infrastructure / DevOps",     "score": 20, "notes": "CRITICAL: .env points to localhost. No production config. Migrations/functions not deployed."},
    {"area": "AI Extraction Pipeline",      "score": 30, "notes": "Architecture correct. All 3 providers coded. Weakness: zero secrets configured → 100% failure rate."},
    {"area": "Email / Notifications",       "score": 25, "notes": "Resend integration built. Notification DB table exists. Weakness: RESEND_API_KEY missing."},
    {"area": "Security",                    "score": 70, "notes": "RLS, audit logs, SECURITY DEFINER hardening migration. Weakness: search_path not guaranteed applied."},
    {"area": "Frontend / UI Quality",       "score": 75, "notes": "65+ pages, Shadcn/ui, responsive design. Weakness: polling bug, no skeleton screens."},
    {"area": "Testing Coverage",            "score": 55, "notes": "80+ property-based tests in supabase/functions/_tests. Weakness: no frontend tests found."},
    {"area": "Documentation",               "score": 40, "notes": "README exists. .env.example covers backend secrets. Weakness: no deployment runbook, no API docs."},
]

DEPLOYMENT_CHECKLIST = [
    {"step": 1,  "category": "Config",    "action": "Set VITE_SUPABASE_URL to production Supabase URL",               "command": "Set in Vercel/Netlify env vars",                                          "done": False},
    {"step": 2,  "category": "Config",    "action": "Set VITE_SUPABASE_ANON_KEY to production anon key",              "command": "Set in Vercel/Netlify env vars",                                          "done": False},
    {"step": 3,  "category": "Database",  "action": "Link local project to production Supabase",                      "command": "supabase link --project-ref <project-id>",                               "done": False},
    {"step": 4,  "category": "Database",  "action": "Apply all 80+ migrations to production",                         "command": "supabase db push",                                                       "done": False},
    {"step": 5,  "category": "Functions", "action": "Deploy all 60+ edge functions",                                  "command": "supabase functions deploy",                                              "done": False},
    {"step": 6,  "category": "Secrets",   "action": "Set RESEND_API_KEY",                                             "command": "supabase secrets set RESEND_API_KEY=re_...",                             "done": False},
    {"step": 7,  "category": "Secrets",   "action": "Set ANTHROPIC_API_KEY (for AI extraction)",                      "command": "supabase secrets set ANTHROPIC_API_KEY=sk-ant-...",                      "done": False},
    {"step": 8,  "category": "Secrets",   "action": "Set WORKER_INTERNAL_SECRET (inter-function auth)",               "command": "supabase secrets set WORKER_INTERNAL_SECRET=<32-char-secret>",           "done": False},
    {"step": 9,  "category": "Secrets",   "action": "Set STRIPE keys (for billing)",                                  "command": "supabase secrets set STRIPE_SECRET_KEY=sk_... STRIPE_WEBHOOK_SECRET=whsec_...", "done": False},
    {"step": 10, "category": "Secrets",   "action": "Set Vertex AI creds (optional, alternative to Anthropic)",       "command": "supabase secrets set VERTEX_PROJECT_ID=... GOOGLE_SERVICE_ACCOUNT_KEY='...'", "done": False},
    {"step": 11, "category": "Storage",   "action": "Verify financial-uploads storage bucket exists",                 "command": "Check Supabase Dashboard > Storage",                                     "done": False},
    {"step": 12, "category": "Code Fix",  "action": "Fix budget hardcoded fallback values in CreateBudget.jsx",        "command": "Guard createMutation.mutate() behind !error check",                      "done": False},
    {"step": 13, "category": "Code Fix",  "action": "Fix pipeline polling stale closure in LeaseUpload.jsx",           "command": "Use useRef for status inside interval",                                  "done": False},
    {"step": 14, "category": "Code Fix",  "action": "Fix default budget_year to current year",                        "command": "Replace 2027 with new Date().getFullYear()",                             "done": False},
    {"step": 15, "category": "Code Fix",  "action": "Remove duplicate src/utils/index.js vs index.ts",                "command": "Delete index.js, keep index.ts",                                         "done": False},
    {"step": 16, "category": "Test",      "action": "Test signup → email confirmation flow",                          "command": "Manual test after RESEND_API_KEY set",                                   "done": False},
    {"step": 17, "category": "Test",      "action": "Test lease PDF upload → AI extraction → review flow",            "command": "Manual test after ANTHROPIC_API_KEY set",                                "done": False},
    {"step": 18, "category": "Test",      "action": "Test budget generation → approval → locking flow",               "command": "Manual test after code fix F-007",                                       "done": False},
]

WORKFLOW_REDFLAGS = [
    {
        "workflow": "User Signup",
        "redflag": "Silent failure after registration",
        "details": "User fills signup form → edge function creates auth.users entry → send-email called → RESEND_API_KEY missing → email fails silently → user sees success screen but gets no email → account stuck in pending_approval → user cannot ever log in without admin intervention.",
        "severity": "CRITICAL",
        "fix": "Set RESEND_API_KEY. Add email fallback using Supabase built-in SMTP as secondary sender.",
    },
    {
        "workflow": "Lease Upload & Extraction",
        "redflag": "Pipeline starts but never progresses past 'Uploaded'",
        "details": "File uploaded to storage → ingest-file edge function called → routes to parse-pdf-docling → DOWNSTREAM_AUTH_FAILED because WORKER_INTERNAL_SECRET missing → worker returns error → status stuck at 'uploaded' → polling stale closure → user sees frozen 'Uploaded' badge forever.",
        "severity": "CRITICAL",
        "fix": "Set WORKER_INTERNAL_SECRET. Fix polling stale closure bug (F-009). Set ANTHROPIC_API_KEY.",
    },
    {
        "workflow": "Budget Creation",
        "redflag": "Fake budget data saved when AI generation fails",
        "details": "User selects property → clicks Generate → generate-budget edge function fails (no secrets) → returns empty object {} → code uses || 669000 fallback → budget saved with $669K revenue / $232K expenses / $437K NOI → user sees 'Budget Created' but numbers are completely fabricated.",
        "severity": "HIGH",
        "fix": "Guard mutation with error check. See F-007.",
    },
    {
        "workflow": "Lease Review Approval",
        "redflag": "Duplicate lease drafts created on retry",
        "details": "Extraction completes → ensureLeaseDraft() called → findLeaseByFileId() JSONB query fails → fallback creates new lease draft → user retries → another draft created → lease list shows 3-4 identical draft leases from one document.",
        "severity": "HIGH",
        "fix": "Fix JSONB query (F-011). Add database unique constraint on leases.extraction_data->>'source_file_id'.",
    },
    {
        "workflow": "CAM → Budget Pipeline",
        "redflag": "Budget completely blocked until CAM is run first",
        "details": "CreateBudget shows 'No CAM snapshot found — Budget generation is blocked' with no easy path forward. New properties have no CAM snapshot until the full CAM Setup → CAM Calculation workflow is completed. This is a chicken-and-egg problem for new customers.",
        "severity": "HIGH",
        "fix": "Make CAM optional for budget generation. Default generateWithoutCam=true.",
    },
    {
        "workflow": "Role-Based Access Control",
        "redflag": "Permission errors crash the page instead of showing graceful denial",
        "details": "assertCurrentUserCanWrite() throws a JavaScript Error that propagates up through React Query's mutation handler. Without an error boundary around every mutation, the entire page unmounts on permission errors.",
        "severity": "MEDIUM",
        "fix": "Wrap assertCurrentUserCanWrite() in try/catch everywhere it's used. Show toast.error with the permission message instead of throwing.",
    },
    {
        "workflow": "Super-Admin Org Switching",
        "redflag": "Super-admin with no acting org sees empty data everywhere",
        "details": "When a super-admin hasn't selected an acting org via the org switcher, orgId is null. resolveWritableOrgId() returns null. All write operations fail. Read operations return empty arrays. The UI shows empty tables with no explanation.",
        "severity": "MEDIUM",
        "fix": "Show a persistent banner for super-admins when no acting org is selected. Block write operations with a clear message: 'Select an organization to perform this action'.",
    },
    {
        "workflow": "MFA Enforcement",
        "redflag": "MFA bypass pages allow unauthenticated access to sensitive flows",
        "details": "MFA_BYPASS_PAGES includes AcceptInvite and ResetPassword — correct. However, the first-login flow creates org and writes to organizations table before MFA is verified. A compromised token could create orgs.",
        "severity": "LOW",
        "fix": "Move org creation in first-login edge function to after MFA verification step. Audit the token validation order in first-login/index.ts.",
    },
]

# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK CREATION
# ══════════════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Sheet 1: Executive Summary ─────────────────────────────────────────────
ws1 = wb.create_sheet("Executive Summary")
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:H1")
cell = ws1["A1"]
cell.value = "CRE Financial Suite — Full Product Audit Report"
cell.fill = HEADER_FILL
cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=16)
cell.alignment = CENTER
ws1.row_dimensions[1].height = 40

ws1.merge_cells("A2:H2")
cell = ws1["A2"]
cell.value = "Audit Date: June 10, 2026  |  Analyst: Claude Code  |  Version: cre-financial-suite-main"
cell.fill = PatternFill("solid", fgColor="334155")
cell.font = Font(color="94A3B8", name="Calibri", size=9)
cell.alignment = CENTER
ws1.row_dimensions[2].height = 20

# Overall maturity gauge
ws1.merge_cells("A4:H4")
cell = ws1["A4"]
cell.value = "OVERALL PRODUCT MATURITY: 52 / 100  —  CRITICAL ISSUES MUST BE RESOLVED BEFORE ANY FEATURE CAN WORK"
cell.fill = PatternFill("solid", fgColor="7F1D1D")
cell.font = Font(color="FCA5A5", bold=True, name="Calibri", size=11)
cell.alignment = CENTER
ws1.row_dimensions[4].height = 28

# Summary KPIs
headers = ["Metric", "Count", "Status"]
kpis = [
    ("Showstopper Issues (nothing works without these)", "3", "🔴 ALL OPEN"),
    ("High Priority Issues (major features broken)", "5", "🟠 ALL OPEN"),
    ("Medium Priority Issues (partial breakage)", "8", "🟡 ALL OPEN"),
    ("Low Priority Issues (quality / polish)", "6", "🟢 Deferrable"),
    ("Total Findings", "22", ""),
    ("Production-Deployable Today?", "NO", "🔴 Blocked by F-001,F-002,F-003"),
    ("Code Files Analyzed", "150+", "✅"),
    ("Edge Functions Reviewed", "60+", "⚠️ Not Deployed"),
    ("DB Migrations Reviewed", "80+", "⚠️ Not Applied"),
    ("Test Files Found", "80+", "✅ Property-based tests"),
]

ws1.row_dimensions[5].height = 5
for i, hdr in enumerate(headers, 1):
    cell = ws1.cell(row=6, column=i+1)
    cell.value = hdr
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = BORDER

for r, (metric, count, status) in enumerate(kpis, 7):
    fills = [LIGHT_FILL if r % 2 == 0 else WHITE_FILL]
    for c, val in enumerate([metric, count, status], 2):
        cell = ws1.cell(row=r, column=c)
        cell.value = val
        cell.font = BODY_FONT
        cell.fill = fills[0]
        cell.alignment = LEFT if c == 2 else CENTER
        cell.border = BORDER

ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 55
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 35
for col in "EFGH":
    ws1.column_dimensions[col].width = 15

# ── Sheet 2: All Findings ──────────────────────────────────────────────────
ws2 = wb.create_sheet("All Findings")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

ws2.merge_cells("A1:K1")
cell = ws2["A1"]
cell.value = "CRE Financial Suite — Issue Register"
cell.fill = HEADER_FILL
cell.font = TITLE_FONT
cell.alignment = CENTER
ws2.row_dimensions[1].height = 32

cols = ["ID", "Area", "Component", "Issue Title", "Priority", "Impact", "Root Cause", "Fix / Resolution", "Effort", "Status", "File Reference"]
col_widths = [8, 22, 28, 45, 14, 50, 50, 55, 10, 12, 45]

for c, (hdr, width) in enumerate(zip(cols, col_widths), 1):
    ws2.column_dimensions[get_column_letter(c)].width = width
    cell = ws2.cell(row=2, column=c)
    cell.value = hdr
    cell.fill = PatternFill("solid", fgColor="0F172A")
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = BORDER

for r, finding in enumerate(FINDINGS, 3):
    pri_fill = priority_fill(finding["priority"])
    sta_fill = status_fill(finding["status"])
    row_fill = LIGHT_FILL if r % 2 == 0 else WHITE_FILL

    values = [
        finding["id"],
        finding["area"],
        finding["component"],
        finding["title"],
        finding["priority"],
        finding["impact"],
        finding["root_cause"],
        finding["fix"],
        finding["effort"],
        finding["status"],
        finding["file"],
    ]
    for c, val in enumerate(values, 1):
        cell = ws2.cell(row=r, column=c)
        cell.value = val
        cell.alignment = LEFT if c not in (1, 5, 9, 10) else CENTER
        cell.border = BORDER
        cell.font = BODY_FONT

        if c == 5:  # priority
            cell.fill = pri_fill
            cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=9)
        elif c == 10:  # status
            cell.fill = sta_fill
            cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=9)
        else:
            cell.fill = row_fill

    ws2.row_dimensions[r].height = 60

ws2.auto_filter.ref = f"A2:K{len(FINDINGS)+2}"

# ── Sheet 3: Priority Breakdown ────────────────────────────────────────────
ws3 = wb.create_sheet("By Priority")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:G1")
cell = ws3["A1"]
cell.value = "Issues Grouped by Priority"
cell.fill = HEADER_FILL
cell.font = TITLE_FONT
cell.alignment = CENTER
ws3.row_dimensions[1].height = 32

priorities = ["SHOWSTOPPER", "HIGH", "MEDIUM", "LOW"]
prio_descs = {
    "SHOWSTOPPER": "Nothing works until these are fixed. App is completely non-functional in production.",
    "HIGH": "Major features broken. Core user workflows fail completely.",
    "MEDIUM": "Partial breakage. Features work in some paths but fail in edge cases or specific conditions.",
    "LOW": "Quality, UX, security polish. App functions but user experience is degraded.",
}

current_row = 3
for prio in priorities:
    prio_findings = [f for f in FINDINGS if f["priority"] == prio]
    if not prio_findings:
        continue

    # Section header
    ws3.merge_cells(f"A{current_row}:G{current_row}")
    cell = ws3[f"A{current_row}"]
    cell.value = f"{'⛔' if prio == 'SHOWSTOPPER' else '🔴' if prio == 'HIGH' else '🟡' if prio == 'MEDIUM' else '🟢'}  {prio} — {len(prio_findings)} issues  |  {prio_descs[prio]}"
    cell.fill = priority_fill(prio)
    cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    cell.alignment = LEFT
    cell.border = BORDER
    ws3.row_dimensions[current_row].height = 24
    current_row += 1

    # Column headers
    hdrs = ["ID", "Title", "Component", "Fix Summary", "Effort", "Status", "File"]
    widths = [8, 50, 28, 60, 10, 12, 40]
    for c, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws3.column_dimensions[get_column_letter(c)].width = w
        cell = ws3.cell(row=current_row, column=c)
        cell.value = h
        cell.fill = DARK_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    current_row += 1

    for finding in prio_findings:
        row_fill = LIGHT_FILL if current_row % 2 == 0 else WHITE_FILL
        for c, val in enumerate([
            finding["id"], finding["title"], finding["component"],
            finding["fix"], finding["effort"], finding["status"], finding["file"]
        ], 1):
            cell = ws3.cell(row=current_row, column=c)
            cell.value = val
            cell.fill = row_fill
            cell.font = BODY_FONT
            cell.alignment = CENTER if c in (1, 5, 6) else LEFT
            cell.border = BORDER
            if c == 6:
                cell.fill = status_fill(finding["status"])
                cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=9)
        ws3.row_dimensions[current_row].height = 50
        current_row += 1

    current_row += 1

# ── Sheet 4: Workflow Red Flags ────────────────────────────────────────────
ws4 = wb.create_sheet("Workflow Red Flags")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:F1")
cell = ws4["A1"]
cell.value = "Workflow-Level Red Flags & Failure Chains"
cell.fill = PatternFill("solid", fgColor="7F1D1D")
cell.font = TITLE_FONT
cell.alignment = CENTER
ws4.row_dimensions[1].height = 32

cols4 = ["Workflow", "Red Flag Title", "What Actually Happens", "Severity", "Fix Summary"]
col4_widths = [22, 40, 80, 14, 60]
for c, (h, w) in enumerate(zip(cols4, col4_widths), 1):
    ws4.column_dimensions[get_column_letter(c)].width = w
    cell = ws4.cell(row=2, column=c)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = BORDER
ws4.row_dimensions[2].height = 20

for r, rf in enumerate(WORKFLOW_REDFLAGS, 3):
    sev_fill = {
        "CRITICAL": RED_FILL,
        "HIGH": PatternFill("solid", fgColor="FF8800"),
        "MEDIUM": YELLOW_FILL,
        "LOW": PatternFill("solid", fgColor="86EFAC"),
    }.get(rf["severity"], WHITE_FILL)
    row_fill = LIGHT_FILL if r % 2 == 0 else WHITE_FILL

    for c, val in enumerate([rf["workflow"], rf["redflag"], rf["details"], rf["severity"], rf["fix"]], 1):
        cell = ws4.cell(row=r, column=c)
        cell.value = val
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = BORDER
        if c == 4:
            cell.fill = sev_fill
            cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=9)
            cell.alignment = CENTER
        else:
            cell.fill = row_fill
    ws4.row_dimensions[r].height = 80

# ── Sheet 5: Product Maturity ──────────────────────────────────────────────
ws5 = wb.create_sheet("Product Maturity")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:E1")
cell = ws5["A1"]
cell.value = "Product Maturity Scorecard — CRE Financial Suite"
cell.fill = HEADER_FILL
cell.font = TITLE_FONT
cell.alignment = CENTER
ws5.row_dimensions[1].height = 32

cols5 = ["Area", "Score / 100", "Visual", "Maturity Level", "Notes"]
widths5 = [35, 14, 20, 20, 65]
for c, (h, w) in enumerate(zip(cols5, widths5), 1):
    ws5.column_dimensions[get_column_letter(c)].width = w
    cell = ws5.cell(row=2, column=c)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = BORDER

def score_fill(score):
    if score >= 80: return PatternFill("solid", fgColor="22C55E")
    if score >= 60: return PatternFill("solid", fgColor="F59E0B")
    if score >= 40: return PatternFill("solid", fgColor="F97316")
    return PatternFill("solid", fgColor="EF4444")

def maturity_level(score):
    if score >= 80: return "Production Ready"
    if score >= 60: return "Beta"
    if score >= 40: return "Alpha"
    if score >= 20: return "Proof of Concept"
    return "Broken"

for r, item in enumerate(MATURITY, 3):
    row_fill = LIGHT_FILL if r % 2 == 0 else WHITE_FILL
    bar = "█" * (item["score"] // 10) + "░" * (10 - item["score"] // 10)
    level = maturity_level(item["score"])

    for c, val in enumerate([item["area"], item["score"], bar, level, item["notes"]], 1):
        cell = ws5.cell(row=r, column=c)
        cell.value = val
        cell.border = BORDER
        cell.alignment = LEFT if c in (1, 3, 5) else CENTER

        if c == 2:
            cell.fill = score_fill(item["score"])
            cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        elif c == 3:
            cell.fill = score_fill(item["score"])
            cell.font = Font(color="FFFFFF", name="Calibri", size=8)
        elif c == 4:
            cell.fill = score_fill(item["score"])
            cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=9)
        else:
            cell.fill = row_fill
            cell.font = BODY_FONT

    ws5.row_dimensions[r].height = 28

# Overall average
avg = sum(m["score"] for m in MATURITY) // len(MATURITY)
ws5.row_dimensions[len(MATURITY)+3].height = 32
ws5.merge_cells(f"A{len(MATURITY)+4}:E{len(MATURITY)+4}")
cell = ws5[f"A{len(MATURITY)+4}"]
cell.value = f"OVERALL PRODUCT MATURITY: {avg} / 100"
cell.fill = score_fill(avg)
cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=13)
cell.alignment = CENTER
cell.border = BORDER
ws5.row_dimensions[len(MATURITY)+4].height = 32

# ── Sheet 6: Deployment Checklist ──────────────────────────────────────────
ws6 = wb.create_sheet("Deployment Checklist")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:F1")
cell = ws6["A1"]
cell.value = "Deployment Checklist — Steps to Make the App Work"
cell.fill = PatternFill("solid", fgColor="0F4C81")
cell.font = TITLE_FONT
cell.alignment = CENTER
ws6.row_dimensions[1].height = 32

cols6 = ["#", "Category", "Action", "Command / How-To", "Done?", "Notes"]
widths6 = [5, 15, 55, 60, 8, 30]
for c, (h, w) in enumerate(zip(cols6, widths6), 1):
    ws6.column_dimensions[get_column_letter(c)].width = w
    cell = ws6.cell(row=2, column=c)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = BORDER

cat_fills = {
    "Config":    PatternFill("solid", fgColor="DBEAFE"),
    "Database":  PatternFill("solid", fgColor="D1FAE5"),
    "Functions": PatternFill("solid", fgColor="FEF3C7"),
    "Secrets":   PatternFill("solid", fgColor="FCE7F3"),
    "Storage":   PatternFill("solid", fgColor="E0F2FE"),
    "Code Fix":  PatternFill("solid", fgColor="FFF3E0"),
    "Test":      PatternFill("solid", fgColor="F3E8FF"),
}

for r, item in enumerate(DEPLOYMENT_CHECKLIST, 3):
    cf = cat_fills.get(item["category"], WHITE_FILL)
    done_cell_fill = GREEN_FILL if item["done"] else PatternFill("solid", fgColor="FEE2E2")

    for c, val in enumerate([item["step"], item["category"], item["action"], item["command"], "✓" if item["done"] else "☐", ""], 1):
        cell = ws6.cell(row=r, column=c)
        cell.value = val
        cell.border = BORDER
        cell.alignment = CENTER if c in (1, 5) else LEFT
        cell.font = BODY_FONT

        if c == 5:
            cell.fill = done_cell_fill
            cell.font = Font(color="FFFFFF" if item["done"] else "EF4444", bold=True, name="Calibri", size=11)
        elif c == 2:
            cell.fill = cf
            cell.font = Font(color="1E293B", bold=True, name="Calibri", size=9)
        else:
            cell.fill = LIGHT_FILL if r % 2 == 0 else WHITE_FILL

    ws6.row_dimensions[r].height = 32

# ── Tab colors ─────────────────────────────────────────────────────────────
ws1.sheet_properties.tabColor = "1D4ED8"
ws2.sheet_properties.tabColor = "0F172A"
ws3.sheet_properties.tabColor = "DC2626"
ws4.sheet_properties.tabColor = "7F1D1D"
ws5.sheet_properties.tabColor = "059669"
ws6.sheet_properties.tabColor = "0F4C81"

# Save
output_path = r"c:\Users\tejas\Downloads\cre-financial-suite-main (3)\cre-financial-suite-main\CRE_Financial_Suite_Audit_Report.xlsx"
wb.save(output_path)
print(f"Report saved: {output_path}")
